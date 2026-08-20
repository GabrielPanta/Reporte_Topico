import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_sample_files():
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. Archivo 1: Datos de Trabajadores (con Apellidos y Nombres en columnas separadas o directas)
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "Trabajadores"
    
    headers1 = [
        "Regimen", "RutTrabajador", "CodigoTrabajador", "Ap.Paterno", "Ap. Materno", "Nombre", 
        "FechaNacimiento", "Sexo", "Edad", "FechaInicioPeriodo", 
        "FechaInicioContrato", "FechaTerminoContrato", "Oficio"
    ]
    ws1.append(headers1)
    
    data1 = [
        ["Agrario", "70112233", "TRAB-001", "Pérez", "Ramos", "Juan Carlos", "1992-04-15", "M", 34, "2026-01-01", "2022-03-15", "2026-12-31", "Cosechador"],
        ["Agrario", "70223344", "TRAB-002", "Rodríguez", "Solís", "KASSANDRA EUFEMIA", "1995-08-22", "F", 31, "2026-01-01", "2021-06-01", "2026-12-31", "Seleccionadora"],
        ["General", "70334455", "TRAB-003", "Sánchez", "Morales", "Carlos Alberto", "1988-11-03", "M", 37, "2026-01-01", "2020-01-10", "Indeterminado", "Supervisor de Campo"],
        ["Agrario", "70445566", "TRAB-004", "Gómez", "Torres", "Ana Lucía", "1998-02-18", "F", 28, "2026-01-01", "2023-08-20", "2026-12-31", "Empacadora"],
        ["Agrario", "70556677", "TRAB-005", "Mendoza", "Castro", "Luis Fernando", "1990-07-30", "M", 36, "2026-01-01", "2019-11-05", "Indeterminado", "Técnico de Riego"],
        ["Agrario", "70667788", "TRAB-006", "Vargas", "Silva", "Patricia Sofía", "1994-09-12", "F", 31, "2026-01-01", "2022-09-12", "2026-12-31", "Evaluadora de Calidad"],
        ["Agrario", "70778899", "TRAB-007", "Alva", "Paredes", "Jorge Luis", "1989-12-05", "M", 36, "2026-01-01", "2021-04-18", "2026-12-31", "Conductor"],
        ["Agrario", "70889900", "TRAB-008", "Fernández", "Quintana", "Rosa María", "1993-03-27", "F", 33, "2026-01-01", "2020-07-22", "Indeterminado", "Fitosanidad"],
        ["Agrario", "70990011", "TRAB-009", "Chávez", "Vega", "Diego Armando", "1996-05-14", "M", 30, "2026-01-01", "2023-02-14", "2026-12-31", "Estibador"],
        ["Agrario", "71001122", "TRAB-010", "Navarro", "Cruz", "Carmen Rosa", "1991-10-08", "F", 34, "2026-01-01", "2022-10-01", "2026-12-31", "Monitor SST"]
    ]
    for row in data1:
        ws1.append(row)

    for col in ws1.iter_cols(min_row=1, max_row=len(data1)+1, min_col=1, max_col=len(headers1)):
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 15)
        for cell in col:
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb1.save("ejemplo_1_trabajadores.xlsx")
    print("Creado: ejemplo_1_trabajadores.xlsx")

    # 2. Archivo 2: Último Día Laborado & Labores (con NOMBRE_ESTACION, HoraInicio, ULTIMO DIA, Tiene Digitacion, etc.)
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Ultimo_Dia_Labores"
    
    headers2 = [
        "RutTrabajador", "Tiene Digitacion (jornal)", "Zona Labores", "SubCentroCosto / Cuartel",
        "ACTIVIDAD", "LABOR", "ENCARGADO", "NOMBRE_ESTACION", "CODIGO BUS", "RUTA", "HoraInicio", "ULTIMO DIA"
    ]
    ws2.append(headers2)
    
    data2 = [
        ["70112233", "SÍ", "Fundo San José", "Cuartel C-12 (Palto)", "COSECHA", "Cosecha de Palta Hass", "Roberto Gómez", "ESTACIÓN PACKING NORTE", "BUS-04", "Ruta 1 - Caserío Central", "06:00 AM", "2026-08-18"],
        ["70223344", "NO", "Planta Packing", "Línea de Empaque 1", "LICENCIA POR MATERNIDAD", "Selección y Calibrado", "Mariela Rojas", "ESTACIÓN PRINCIPAL", "-", "-", "07:00 AM", "2026-10-15"],
        ["70334455", "SÍ", "Fundo San José", "Sector A General", "SUPERVISIÓN", "Control de Cuadrillas", "Carlos Sánchez", "ESTACIÓN FUNDO CENTRAL", "CAM-01", "Ruta 2 - Sector Norte", "05:30 AM", "2026-08-18"],
        ["70445566", "NO", "Planta Packing", "Área Terminado", "PERMISO CON GOCE", "Envasado y Pesaje", "Mariela Rojas", "ESTACIÓN PACKING SUR", "-", "-", "07:00 AM", "2026-08-25"],
        ["70556677", "NO", "Fundo San José", "Estación de Riego 2", "VACACIONES", "Control de Presurizado", "Hernán Silva", "ESTACIÓN RIEGO", "-", "-", "06:00 AM", "2026-08-30"],
        ["70667788", "SÍ", "Fundo Santa Rosa", "Cuartel B-04 (Arándano)", "EVALUACIÓN", "Muestreo de Brix", "Patricia Vargas", "ESTACIÓN SANTA ROSA", "BUS-02", "Ruta 3 - Los Olivos", "06:00 AM", "2026-08-18"],
        ["70778899", "NO", "Logística Central", "Flota Vehicular", "PERMISO PARTICULAR", "Traslado de Cosecha", "Esteban Quispe", "ESTACIÓN COCHERA", "-", "-", "06:00 AM", "2026-08-28"],
        ["70889900", "SÍ", "Fundo San José", "Cuartel D-08 (Uva)", "APLICACIÓN", "Evaluación Fitosanitaria", "Guillermo Paz", "ESTACIÓN SAN JOSÉ", "BUS-05", "Ruta 1 - Caserío Central", "05:45 AM", "2026-08-18"],
        ["70990011", "NO", "Planta Packing", "Cámara Fría 1", "DESCANSO MÉDICO", "Paletizado y Despacho", "Manuel Farfán", "ESTACIÓN PACKING NORTE", "-", "-", "02:00 PM", "2026-08-22"],
        ["71001122", "SÍ", "Todas las Sedes", "SST General", "INSPECCIÓN", "Charla 5 Minutos y Ronda SST", "Carmen Navarro", "ESTACIÓN SST", "BUS-01", "Ruta Expresa", "05:30 AM", "2026-08-18"]
    ]
    for row in data2:
        ws2.append(row)

    for col in ws2.iter_cols(min_row=1, max_row=len(data2)+1, min_col=1, max_col=len(headers2)):
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 16)
        for cell in col:
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb2.save("ejemplo_2_ultimo_dia_labores.xlsx")
    print("Creado: ejemplo_2_ultimo_dia_labores.xlsx")

    # 3. Archivo 3: Marcaciones
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.title = "Marcaciones"
    
    headers3 = ["Documento", "FechaHora", "Puerta", "NOMBRE_ESTACION", "TIPO_ESTACION"]
    ws3.append(headers3)
    
    data3 = [
        ["70112233", "2026-08-18 05:45:12", "Puerta Principal Fundo", "BUS-04", "BUS"],
        ["70334455", "2026-08-18 05:30:10", "Puerta Administrativa", "CAM-01", "BUS"],
        ["70667788", "2026-08-18 05:50:22", "Puerta Packing", "BUS-02", "BUS"],
        ["70889900", "2026-08-18 05:40:05", "Puerta Principal Fundo", "BUS-05", "BUS"],
        ["71001122", "2026-08-18 05:35:18", "Puerta Principal Fundo", "OFICINA SST", "FIJA"]
    ]
    for row in data3:
        ws3.append(row)

    for col in ws3.iter_cols(min_row=1, max_row=len(data3)+1, min_col=1, max_col=len(headers3)):
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws3.column_dimensions[col_letter].width = max(max_len + 4, 18)
        for cell in col:
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb3.save("ejemplo_3_marcaciones.xlsx")
    print("Creado: ejemplo_3_marcaciones.xlsx")

    # 4. Archivo 4: Catálogo de Buses y Rutas
    wb4 = openpyxl.Workbook()
    ws4 = wb4.active
    ws4.title = "Buses_y_Rutas"

    headers4 = [
        "Predio", "Transportista", "Codigo Campo", "Patente", "Descripcion Ruta",
        "Troncal", "Cuartel", "Observaciones", "Capacidad", "Cant. Personas",
        "ASIENTOS DISPONIBLES", "COSTO", "Bonificacion", "Total", "RESUMEN VIAJES"
    ]
    ws4.append(headers4)

    data4 = [
        ["Fundo San José", "Transportes del Norte S.A.C.", "BUS-04", "BUS-04", "Ruta 1 - Caserío Central", "Troncal A", "C-12", "Operativo", 45, 42, 3, 180, 0, 180, "Turno Mañana"],
        ["Planta Packing", "Servicios Verfrut", "CAM-01", "CAM-01", "Ruta 2 - Sector Norte", "Troncal B", "A-General", "Operativo", 30, 28, 2, 150, 0, 150, "Turno Mañana"],
        ["Fundo Santa Rosa", "Transportes del Norte S.A.C.", "BUS-02", "BUS-02", "Ruta 3 - Los Olivos", "Troncal C", "B-04", "Operativo", 50, 48, 2, 200, 0, 200, "Turno Mañana"],
        ["Fundo San José", "Transportes Rápidos", "BUS-05", "BUS-05", "Ruta 1 - Caserío Central", "Troncal A", "D-08", "Operativo", 45, 44, 1, 180, 0, 180, "Turno Mañana"]
    ]
    for row in data4:
        ws4.append(row)

    for col in ws4.iter_cols(min_row=1, max_row=len(data4)+1, min_col=1, max_col=len(headers4)):
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws4.column_dimensions[col_letter].width = max(max_len + 4, 16)
        for cell in col:
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb4.save("ejemplo_4_buses_rutas.xlsx")
    print("Creado: ejemplo_4_buses_rutas.xlsx")

    # 5. Archivo 5: Catálogo de Cuadrillas & Encargados
    wb5 = openpyxl.Workbook()
    ws5 = wb5.active
    ws5.title = "Cuadrillas"

    headers5 = [
        "IDCUADRILLA", "Codigo Encargado", "Nombre Encargado", "Descripcion",
        "RUT/DNI Encargado", "Fecha Ingreso Encargado", "Codigo Trabajador"
    ]
    ws5.append(headers5)

    data5 = [
        [1, "000001", "CESAR ALBERTO MELGAR MARCHAN", "MELGAR MARCHAN CESAR ALBERTO", "43491155", "2021-05-10", "-"],
        [2, "000051", "RICARDO ORLANDO RAMOS LOZADA", "RAMOS LOZADA RICARDO ORLANDO", "41865225", "2020-03-15", "-"],
        [5, "000111", "CARLOS SILUPU ABAD", "SILUPU ABAD CARLOS", "42370580", "2022-08-20", "-"],
        [7, "000055", "HENRY IPANAQUE VIERA", "IPANAQUE VIERA HENRY", "44040239", "2020-01-12", "-"],
        [202, "000202", "JOSE ADALBERTO SUAREZ MAZA", "SUAREZ MAZA JOSE ADALBERTO", "45211189", "2019-11-04", "-"]
    ]
    for row in data5:
        ws5.append(row)

    for col in ws5.iter_cols(min_row=1, max_row=len(data5)+1, min_col=1, max_col=len(headers5)):
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws5.column_dimensions[col_letter].width = max(max_len + 4, 18)
        for cell in col:
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb5.save("ejemplo_5_cuadrillas.xlsx")
    print("Creado: ejemplo_5_cuadrillas.xlsx")

if __name__ == "__main__":
    create_sample_files()
